#!/usr/bin/env python3
"""Convert the glavoy expense spreadsheets into CSVs the app can import.

Usage:
    python import_xlsx.py --raw C:/temp/raw_data.xlsx --y2026 C:/temp/glavoy2026.xlsx --out ./import_csv

Produces in --out:
    transactions.csv  daily expenses + monthly income, booked against the
                      hidden "Imported history" account (acc-history)
    fx_rates.csv      daily UGX/USD/CAD rates from the xrates sheet
    validation.txt    monthly totals per category, to eyeball against the
                      'monthly ex' sheet

Notes
-----
* Daily expenses are taken from raw_data.xlsx (2006 -> its last row) and then
  from the 2026 workbook's 'daily ex' sheet for any dates AFTER raw_data ends.
  glavoy2025.xlsx is not needed: 2025 is already inside raw_data.xlsx.
* All ids are deterministic, so re-running the script and re-importing the
  CSVs is safe (rows overwrite themselves, never duplicate).
* Mobile money / Balance sheets are intentionally NOT imported as
  transactions: their spending already exists in 'daily ex' and importing
  both would double-count. Set each real account's current balance as its
  opening balance in the app (Settings > Accounts) when you start.
"""

import argparse
import csv
import datetime as dt
from collections import defaultdict
from pathlib import Path

import openpyxl

HISTORY_ACCOUNT = "acc-history"
INCOME_SALARY_CATEGORY = "cat-income-salary"

# Category column headers exactly as they appear in the sheets, mapped to the
# app's seeded category ids (see lib/data/seed.dart: 'cat-expense-<lowercase>').
EXPENSE_HEADERS = [
    "food", "food_r", "beer", "beer_r", "house", "petrol", "car",
    "motorcycle", "health", "clothes", "recreation", "jose", "misc", "kids",
    "airtime", "bigticket", "water", "rent", "electricity", "Internet",
    "guard", "Dog", "Buziga", "Munyonyo", "DSTV", "worker",
]


def category_id(header: str) -> str:
    return f"cat-expense-{header.lower().replace(' ', '-')}"


def day_str(d: dt.date) -> str:
    return d.strftime("%Y-%m-%d")


def parse_daily_sheet(ws, after: dt.date | None):
    """Yield (date, header, amount, day_note) for non-zero expense cells."""
    headers = [c.value for c in ws[1]]
    col_of = {}
    for idx, h in enumerate(headers):
        if isinstance(h, str) and h.strip() in EXPENSE_HEADERS:
            # First occurrence wins ('beer' appears twice in some sheets;
            # the trailing duplicate is a helper column, not data).
            col_of.setdefault(h.strip(), idx)
    note_col = next(
        (i for i, h in enumerate(headers) if isinstance(h, str) and h.strip().lower() == "notes"),
        None,
    )
    date_col = next(
        i for i, h in enumerate(headers) if isinstance(h, str) and h.strip().lower() == "date"
    )

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_date = row[date_col] if date_col < len(row) else None
        if not isinstance(raw_date, (dt.date, dt.datetime)):
            continue
        date = raw_date.date() if isinstance(raw_date, dt.datetime) else raw_date
        if after is not None and date <= after:
            continue
        if date > dt.date.today():
            continue
        note = None
        if note_col is not None and note_col < len(row):
            v = row[note_col]
            if isinstance(v, str) and v.strip():
                note = v.strip()
        for header, idx in col_of.items():
            value = row[idx] if idx < len(row) else None
            if isinstance(value, (int, float)) and value > 0:
                yield date, header, float(value), note


def build_transactions(raw_path: Path, y2026_path: Path | None):
    rows = []
    monthly_totals = defaultdict(lambda: defaultdict(float))

    wb = openpyxl.load_workbook(raw_path, data_only=True, read_only=True)
    last_date = None
    if "daily_exp_ugx" in wb.sheetnames:
        day_notes = {}
        for date, header, amount, note in parse_daily_sheet(wb["daily_exp_ugx"], None):
            last_date = max(last_date, date) if last_date else date
            rows.append(_expense_row(date, header, amount))
            monthly_totals[date.strftime("%Y-%m")][header] += amount
            if note:
                day_notes[date] = note
        _attach_day_notes(rows, day_notes)

    if "monthly_income" in wb.sheetnames:
        for r in wb["monthly_income"].iter_rows(min_row=2, values_only=True):
            raw_month, _usd, _cad, ugx = (list(r) + [None] * 4)[:4]
            if not isinstance(raw_month, (dt.date, dt.datetime)) or not isinstance(ugx, (int, float)) or ugx <= 0:
                continue
            month = raw_month.date() if isinstance(raw_month, dt.datetime) else raw_month
            rows.append({
                "id": f"imp-inc-{month:%Y%m}",
                "date": day_str(month),
                "kind": "income",
                "amount": round(float(ugx), 2),
                "account_id": HISTORY_ACCOUNT,
                "category_id": INCOME_SALARY_CATEGORY,
                "to_account_id": "",
                "to_amount": "",
                "note": "",
            })
    wb.close()

    if y2026_path is not None and y2026_path.exists():
        wb26 = openpyxl.load_workbook(y2026_path, data_only=True, read_only=True)
        if "daily ex" in wb26.sheetnames:
            day_notes = {}
            new_rows = []
            for date, header, amount, note in parse_daily_sheet(wb26["daily ex"], last_date):
                new_rows.append(_expense_row(date, header, amount))
                monthly_totals[date.strftime("%Y-%m")][header] += amount
                if note:
                    day_notes[date] = note
            _attach_day_notes(new_rows, day_notes)
            rows.extend(new_rows)
        wb26.close()

    return rows, monthly_totals


def _expense_row(date, header, amount):
    return {
        "id": f"imp-{date:%Y%m%d}-{header.lower()}",
        "date": day_str(date),
        "kind": "expense",
        "amount": round(amount, 2),
        "account_id": HISTORY_ACCOUNT,
        "category_id": category_id(header),
        "to_account_id": "",
        "to_amount": "",
        "note": "",
    }


def _attach_day_notes(rows, day_notes):
    """Attach the day's Notes text to that day's 'misc' entry when present,
    otherwise to the first entry of the day."""
    by_date = defaultdict(list)
    for row in rows:
        by_date[row["date"]].append(row)
    for date, note in day_notes.items():
        day_rows = by_date.get(day_str(date))
        if not day_rows:
            continue
        target = next((r for r in day_rows if r["category_id"] == category_id("misc")), day_rows[0])
        target["note"] = note


def build_fx_rates(raw_path: Path):
    wb = openpyxl.load_workbook(raw_path, data_only=True, read_only=True)
    rows = []
    if "xrates" in wb.sheetnames:
        for r in wb["xrates"].iter_rows(min_row=2, values_only=True):
            raw_date, cadugx, cadusd, usdcad, usdugx = (list(r) + [None] * 5)[:5]
            if not isinstance(raw_date, (dt.date, dt.datetime)):
                continue
            date = raw_date.date() if isinstance(raw_date, dt.datetime) else raw_date
            rows.append({
                "id": f"fx-{date:%Y%m%d}",
                "date": day_str(date),
                "usd_ugx": float(usdugx) if isinstance(usdugx, (int, float)) else "",
                "cad_ugx": float(cadugx) if isinstance(cadugx, (int, float)) else "",
                "usd_cad": float(usdcad) if isinstance(usdcad, (int, float)) else "",
                "source": "import",
            })
    wb.close()
    return rows


def write_csv(path: Path, rows, fieldnames):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--raw", required=True, type=Path, help="raw_data.xlsx")
    ap.add_argument("--y2026", type=Path, help="glavoy2026.xlsx (for dates after raw_data ends)")
    ap.add_argument("--out", type=Path, default=Path("import_csv"))
    args = ap.parse_args()

    args.out.mkdir(parents=True, exist_ok=True)

    transactions, monthly_totals = build_transactions(args.raw, args.y2026)
    write_csv(
        args.out / "transactions.csv",
        transactions,
        ["id", "date", "kind", "amount", "account_id", "category_id",
         "to_account_id", "to_amount", "note"],
    )
    print(f"transactions.csv: {len(transactions)} rows")

    fx = build_fx_rates(args.raw)
    write_csv(
        args.out / "fx_rates.csv",
        fx,
        ["id", "date", "usd_ugx", "cad_ugx", "usd_cad", "source"],
    )
    print(f"fx_rates.csv: {len(fx)} rows")

    # Validation report: monthly totals per category (UGX), for comparing
    # against the 'monthly ex' worksheet.
    with (args.out / "validation.txt").open("w", encoding="utf-8") as f:
        for month in sorted(monthly_totals):
            cats = monthly_totals[month]
            total = sum(cats.values())
            f.write(f"{month}  total={total:,.0f}\n")
            for header in EXPENSE_HEADERS:
                if cats.get(header):
                    f.write(f"    {header:<12} {cats[header]:>15,.0f}\n")
    print(f"validation.txt written. Spot-check a few months against 'monthly ex'.")


if __name__ == "__main__":
    main()
